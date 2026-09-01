import os
import glob
import pandas as pd
from openpyxl import load_workbook

# ==============================
# Configuration: Paths & Sheet Names
# ==============================

# Path to the shared Excel file
excel_path = "archive/exp1/exp1.xlsx"

# Define root paths and corresponding sheet names
root_sheet_mapping = {
    "../beagle/analysis/HLA/": "HLA_Beagle",
    "../beagle/analysis/LOS_chr22/": "LOS_Beagle",
    "../beagle/analysis/1KGP_chr22/": "1KGP_Beagle",
    "archive/exp1/HLA_chr6_ALL_seg-1_overlap0/analysis/": "HLA_SCDA",
    "archive/exp1/HLA_chr6_ALL_seg128_overlap8/analysis/": "HLA_Unet",
    "archive/exp1/1KGP_chr22_EUR_seg-1_overlap0/analysis/": "1KGP_SCDA",
    "archive/exp1/LOS_chr22_ALL_seg-1_overlap0/analysis/": "LOS_SCDA",
    "archive/exp1/LOS_chr22_ALL_seg128_overlap8/analysis/": "LOS_Unet"
}

# List of random states
random_states = [1024, 42, 0]

# Load the shared Excel file
wb = load_workbook(excel_path)

for root_path, sheet_name in root_sheet_mapping.items():
    print(f"Processing: {sheet_name} from {root_path}")
    
    # Find all relevant CSV files for this root path
    csv_files = []
    for state in random_states:
        match = glob.glob(os.path.join(root_path, f"LONI_*_{state}.csv"))
        if match:
            csv_files.append(match[0])  # Take the first match (assuming one per state)
        else:
            print(f"Warning: No file found for state {state} in {root_path}")

    # Skip this sheet if no CSV files were found
    if not csv_files:
        print(f"Skipping {sheet_name} (No CSV files found).")
        continue
    
    # Load all CSV files, replacing blank cells with zero
    dfs = [pd.read_csv(file).fillna(0) for file in csv_files]
    
    # Compute mean for numeric columns while keeping MAF_bin unchanged
    numeric_columns = dfs[0].select_dtypes(include=["number"]).columns[1:13]
    mean_df = sum(df[numeric_columns] for df in dfs) / len(dfs)
    mean_df.insert(0, "MAF_bin", dfs[0]["MAF_bin"].values)
    mean_df.insert(1, "Num_SNPs", dfs[0]["Num_SNPs"].values)
    
    # Check if Ignored Loci columns exist
    has_ignored_loci = "Ignored_Loci_Count" in dfs[0].columns and "Ignored_Loci_Percentage" in dfs[0].columns
    if has_ignored_loci:
        ignored_loci_df = dfs[0][["Ignored_Loci_Count", "Ignored_Loci_Percentage"]].copy()
    
    # Open the corresponding sheet in the Excel file
    ws = wb[sheet_name]
    
    # Extract existing missing ratio section headers to determine start positions dynamically
    excel_data = [row for row in ws.iter_rows(values_only=True)]
    excel_df = pd.DataFrame(excel_data)
    
    missing_ratios = ["5% Missing", "15% Missing", "25% Missing", "50% Missing"]
    start_rows = []
    
    for ratio in missing_ratios:
        for idx, value in enumerate(excel_df[0]):
            if isinstance(value, str) and ratio in value:
                start_rows.append(idx + 2)  # Insert data starting from the next row (skip merged header)
                break
    
    for idx, start_row in enumerate(start_rows):
        data_rows = mean_df.iloc[idx*8:(idx+1)*8, :]
        actual_data_rows = len(data_rows)
        end_row = start_row + actual_data_rows - 1
        overall_row = end_row + 1
        
        # Adjust row count if needed
        excel_section_rows = end_row - start_row + 1
        row_diff = actual_data_rows - excel_section_rows
        
        if row_diff < 0:
            for _ in range(abs(row_diff)):
                ws.delete_rows(end_row)
                end_row -= 1
        elif row_diff > 0:
            for _ in range(row_diff):
                ws.insert_rows(end_row + 1)
                end_row += 1
        
        # Insert data block (starting from column B to avoid merged cells in column A)
        for i, row_num in enumerate(range(start_row, end_row + 1)):
            for j, col_num in enumerate(range(2, 10)):  # Start from column B (index 2)
                ws.cell(row=row_num, column=col_num, value=mean_df.iloc[idx*8 + i, j])
        
        # Compute sum of Num_SNPs column and insert it next to Overall label
        total_snps = mean_df.iloc[idx*8:(idx+1)*8, 1].sum()
        ws.cell(row=overall_row, column=3, value=total_snps)  # Insert at column C (next to Overall label)
        
        # Insert overall row (starting from column D)
        for j, col_num in enumerate(range(4, 10)):
            ws.cell(row=overall_row, column=col_num, value=mean_df.iloc[idx*8, j+8])
        
        # Copy ignored loci columns (J and K) from the original CSV instead of averaging
        if has_ignored_loci:
            for i, row_num in enumerate(range(start_row, end_row + 1)):
                ws.cell(row=row_num, column=10, value=ignored_loci_df.iloc[idx*8 + i, 0])
                ws.cell(row=row_num, column=11, value=ignored_loci_df.iloc[idx*8 + i, 1])
    
    print(f"✅ Data inserted into: {sheet_name}")

# Save the updated Excel file
wb.save(excel_path)
print(f"🎯 All sheets updated successfully in: {excel_path}")
 
